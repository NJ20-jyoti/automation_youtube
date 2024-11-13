# import pickle
# import time
# import openpyxl
# import random
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from concurrent.futures import ThreadPoolExecutor
# import undetected_chromedriver as uc
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
# def create_driver(port):
#     """Creates and returns a Chrome driver connected to an existing instance."""
#     options = uc.ChromeOptions()
#     options.add_argument("--disable-infobars")
#     options.add_argument("--disable-extensions")
#     options.add_argument("--window-size=800,600")  # Adjust window size as necessary
#     return uc.Chrome(options=options)

# def load_cookies(driver, cookie_file):
#     """Loads cookies from a pickle file into the provided Selenium driver."""
#     try:
#         driver.get("https://www.youtube.com")  # Load the page first
#         time.sleep(2)  # Allow time for the page to load
#         with open(cookie_file, 'rb') as file:
#             cookies = pickle.load(file)
#             for cookie in cookies:
#                 if 'domain' in cookie and 'youtube.com' in cookie['domain']:
#                     driver.add_cookie(cookie)
#         print(f"Cookies loaded from {cookie_file}")
#         return True
#     except Exception as e:
#         print(f"Error loading cookies from {cookie_file}: {e}")
#         return False

# def post_comment(driver, video_url, comments):
#     """Posts a random comment on the given video URL."""
#     driver.get(video_url)
#     time.sleep(7)  # Wait for the video page to load

#     try:
#         # Attempt to play the video
#         play_button = driver.find_element(By.CSS_SELECTOR, '#movie_player > div.ytp-chrome-bottom > div.ytp-chrome-controls > div.ytp-left-controls > button')
#         play_button.click()
#         time.sleep(3)
#     except Exception as e:
#         print("Could not play the video: ", e)

#     # Scroll down to the comments section
#     driver.execute_script("window.scrollTo(0, 600);")
#     time.sleep(1)

#     try:
#         # Wait for the comment box to be present and click on it
#         WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "ytd-comments ytd-comment-simplebox-renderer")))
#         driver.find_element(By.CSS_SELECTOR, "ytd-comments ytd-comment-simplebox-renderer div#placeholder-area").click()

#         # Select a random comment to post
#         comment = random.choice(comments)
#         driver.find_element(By.CSS_SELECTOR, "#contenteditable-root").send_keys(comment)
#         time.sleep(2)

#         # Click on the submit button
#         driver.find_element(By.ID, "submit-button").click()
#         print(f"Posted comment: {comment}")
#         time.sleep(4)  # Adjust based on YouTube response time
#     except Exception as e:
#         print(f"An error occurred while posting comment: {e}")

# def load_comments_from_excel(file_path):
#     """Loads comments from an Excel file and returns them as a list."""
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     comments = [row[0] for row in sheet.iter_rows(values_only=True) if row[0]]
#     return comments
# def run_for_account(cookie_file, video_url, comments, num_comments):
#     """
#     Run the Selenium operations for a specific account using the specified cookie file.
    
#     Args:
#         cookie_file (str): Path to the cookie file.
#         video_url (str): URL of the video to comment on.
#         comments (list): List of comments to choose from.
#         num_comments (int): Number of comments to post.
#     """
#     # Extract port number from the cookie file name
#     port_str = cookie_file.split('_')[-1].split('.')[0]  # e.g., "port_9222"
#     port = int(port_str.split('_')[1])  # Extract the port number

#     # Create the driver using the specified port
#     driver = create_driver(port)
    
#     try:
#         # Load cookies into the driver
#         if load_cookies(driver, cookie_file):
#             # Post the specified number of comments
#             for _ in range(num_comments):
#                 post_comment(driver, video_url, comments)
#     except Exception as e:
#         print(f"An error occurred while running for account {cookie_file}: {e}")
#     finally:
#         driver.quit()
# def main():
#     accounts = [    
#         "cookies_account_1_port_9222.pkl",
#         "cookies_account_3_port_9224.pkl",
#         "cookies_account_4_port_9225.pkl",
#         "cookies_account_5_port_9226.pkl",
#         "cookies_account_6_port_9227.pkl",
#         "cookies_account_7_port_9228.pkl",
#         "cookies_account_8_port_9229.pkl",
#         "cookies_account_9_port_9230.pkl",
#         "cookies_account_10_port_9231.pkl",
#         "cookies_account_11_port_9232.pkl",
#         "cookies_account_12_port_9233.pkl",
#         "cookies_account_13_port_9234.pkl",
#         "cookies_account_14_port_9235.pkl",
#         "cookies_account_15_port_9236.pkl",
#         "cookies_account_16_port_9237.pkl",
#         "cookies_account_17_port_9238.pkl",
#         "cookies_account_18_port_9239.pkl",
#         "cookies_account_19_port_9240.pkl",
#         "cookies_account_21_port_9242.pkl",
#         "cookies_account_22_port_9243.pkl",
#         "cookies_account_23_port_9244.pkl",
#         "cookies_account_24_port_9245.pkl",
#         "cookies_account_25_port_9246.pkl",
#         "cookies_account_26_port_9247.pkl",
#         "cookies_account_27_port_9248.pkl",
#         "cookies_account_28_port_9249.pkl",
#         "cookies_account_29_port_9250.pkl",
#         "cookies_account_30_port_9251.pkl",
#         "cookies_account_31_port_9252.pkl",
#         "cookies_account_33_port_9254.pkl",
#         "cookies_account_34_port_9255.pkl",
#         "cookies_account_35_port_9256.pkl",
#         "cookies_account_36_port_9257.pkl",
#         "cookies_account_37_port_9258.pkl",
#         "cookies_account_38_port_9259.pkl",
#         "cookies_account_39_port_9260.pkl",
#         "cookies_account_40_port_9261.pkl",
#         "cookies_account_41_port_9262.pkl",
#         "cookies_account_42_port_9263.pkl",
#         "cookies_account_43_port_9264.pkl",
#         "cookies_account_44_port_9265.pkl",
#         "cookies_account_45_port_9266.pkl",
#         "cookies_account_46_port_9267.pkl",
#         "cookies_account_47_port_9268.pkl",
#         "cookies_account_48_port_9269.pkl",
#         "cookies_account_49_port_9270.pkl",
#         "cookies_account_50_port_9271.pkl"
#     ]

    
#     video_url = "https://youtu.be/LZbGkTQ5UTA?si=2bYohE3f7L3Q6w1s"  # Video URL
#     comments_file_path = "100.xlsx"  # Path to your comments Excel file

#     comments = load_comments_from_excel(comments_file_path)

    
#     num_comments= 80
#      # Total iterations to run
#     # max_workers = 3
#      # Number of parallel browser sessions

#     with ThreadPoolExecutor(max_workers=2) as executor:
#         for  cookie_file in accounts:
#             executor.submit(run_for_account,  cookie_file, video_url, comments, num_comments)
# if __name__ == "__main__":
#     main()
import os
import re
import pickle
import time
import pandas as pd
import openpyxl
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor
import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def extract_port_from_filename(filename):
    """Extract the port number from the given filename."""
    match = re.search(r'port_(\d+)', filename)
    if match:
        return int(match.group(1))
    else:
        raise ValueError(f"Could not extract port from filename: {filename}")
def create_driver(port):
    """Creates and returns a Chrome driver connected to an existing instance."""
    options = uc.ChromeOptions()
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    options.add_argument("--window-size=800,600")  # Adjust window size as necessary
    options.add_argument(f"--remote-debugging-port={port}")  # Connect to existing Chrome instance via port
    return uc.Chrome(options=options)


def load_cookies(driver, cookie_file):
    """Loads cookies from a pickle file into the provided Selenium driver."""
    try:
        driver.get("https://www.youtube.com")  # Load the page first
        time.sleep(2)  # Allow time for the page to load
        with open(cookie_file, 'rb') as file:
            cookies = pickle.load(file)
            for cookie in cookies:
                if 'domain' in cookie and 'youtube.com' in cookie['domain']:
                    driver.add_cookie(cookie)
        print(f"Cookies loaded from {cookie_file}")
        return True
    except Exception as e:
        print(f"Error loading cookies from {cookie_file}: {e}")
        return False


def post_comment(driver, video_url, comments):
    """Posts a random comment on the given video URL."""
    driver.get(video_url)
    time.sleep(7)  # Wait for the video page to load

    try:
        # Attempt to play the video
        play_button = driver.find_element(By.CSS_SELECTOR, '#movie_player > div.ytp-chrome-bottom > div.ytp-chrome-controls > div.ytp-left-controls > button')
        play_button.click()
        time.sleep(3)
    except Exception as e:
        print("Could not play the video:", e)

    # Scroll down to the comments section
    driver.execute_script("window.scrollTo(0, 600);")
    time.sleep(1)
     
    try:
        # Wait for the comment box to be present and click on it
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "ytd-comments ytd-comment-simplebox-renderer")))
        driver.find_element(By.CSS_SELECTOR, "ytd-comments ytd-comment-simplebox-renderer div#placeholder-area").click()

        # Select a random comment to post
        comment = random.choice(comments)
        driver.find_element(By.CSS_SELECTOR, "#contenteditable-root").send_keys(comment)
        time.sleep(2)

        # Click on the submit button
        driver.find_element(By.ID, "submit-button").click()
        print(f"Posted comment: {comment}")
        time.sleep(4)  # Adjust based on YouTube response time
    except Exception as e:
        print(f"An error occurred while posting comment: {e}")

def load_comments_from_excel(file_path):
    """Loads comments from an Excel file and returns them as a list."""
    comments = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        # Load full paragraphs from the first column of each row
        comments = [str(row[0]).strip() for row in sheet.iter_rows(values_only=True) if row[0]]
        print(f"Loaded {len(comments)} comments from Excel.")
    except Exception as e:
        print(f"Error loading comments from Excel file: {e}")
    return comments

def run_for_account(cookie_file, video_url, comments, num_comments):
    """Run the Selenium operations for a specific account using the specified cookie file."""
    driver = None  # Initialize driver variable
    try:
        # Extract the port number from the cookie filename
        port = extract_port_from_filename(cookie_file)
        print(f"Using port: {port}")

        driver = create_driver(port)  # Create the Selenium WebDriver

        # Load cookies into the driver
        if load_cookies(driver, cookie_file):
            if not comments:
                print("No comments were loaded. Please check the Excel file.")
                return
             # Post the selected comment
            for _ in range(num_comments):
                # comment = random.choice(comments)  # Choose a random comment
                post_comment(driver, video_url, comments)  # Post the selected comment
                time.sleep(2)
        else:
            print(f"Failed to load cookies for {cookie_file}.")
    
    except Exception as e:
        print(f"An error occurred while running for account {cookie_file}: {e}")
    finally:
        if driver:  # Check if driver was initialized before quitting
            driver.quit()
def main():
    accounts = [
        "cookies_account_1_port_9222.pkl",
        "cookies_account_3_port_9224.pkl",
        "cookies_account_4_port_9225.pkl",
        "cookies_account_5_port_9226.pkl",
     
        "cookies_account_8_port_9229.pkl",
        "cookies_account_9_port_9230.pkl",
        "cookies_account_10_port_9231.pkl",
        "cookies_account_11_port_9232.pkl",
        "cookies_account_12_port_9233.pkl",
        "cookies_account_13_port_9234.pkl",
        "cookies_account_14_port_9235.pkl",
        "cookies_account_15_port_9236.pkl",
        "cookies_account_16_port_9237.pkl",
        "cookies_account_17_port_9238.pkl",
        "cookies_account_18_port_9239.pkl",
        "cookies_account_19_port_9240.pkl",
        "cookies_account_21_port_9242.pkl",
        "cookies_account_22_port_9243.pkl",
        "cookies_account_23_port_9244.pkl",
        "cookies_account_24_port_9245.pkl",
        "cookies_account_25_port_9246.pkl",
        "cookies_account_26_port_9247.pkl",
        "cookies_account_27_port_9248.pkl",
        "cookies_account_28_port_9249.pkl",
        "cookies_account_29_port_9250.pkl",
        "cookies_account_30_port_9251.pkl",
        "cookies_account_31_port_9252.pkl",
        "cookies_account_33_port_9254.pkl",
        "cookies_account_34_port_9255.pkl",
        "cookies_account_35_port_9256.pkl",
        "cookies_account_36_port_9257.pkl",
        "cookies_account_37_port_9258.pkl",
        "cookies_account_38_port_9259.pkl",
        "cookies_account_39_port_9260.pkl",
        "cookies_account_40_port_9261.pkl",
        "cookies_account_41_port_9262.pkl",
        "cookies_account_42_port_9263.pkl",
        "cookies_account_43_port_9264.pkl",
        "cookies_account_44_port_9265.pkl",
        "cookies_account_45_port_9266.pkl",
        "cookies_account_46_port_9267.pkl",
        "cookies_account_47_port_9268.pkl",
        "cookies_account_48_port_9269.pkl",
        "cookies_account_49_port_9270.pkl",
        "cookies_account_50_port_9271.pkl"
    ]

    video_url = "https://youtu.be/LZbGkTQ5UTA?si=2bYohE3f7L3Q6w1s"  # Video URL
    comments_file_path = "100.xlsx"  # Path to your comments Excel file

    comments = load_comments_from_excel(comments_file_path)

    num_comments = 8  # Total number of comments to post

    # Using ThreadPoolExecutor to handle multiple accounts in parallel
    with ThreadPoolExecutor(max_workers=4) as executor:
        for cookie_file in accounts:
            executor.submit(run_for_account, cookie_file, video_url, comments, num_comments)


if __name__ == "__main__":
    main()
